import openpyxl
from openpyxl.utils import get_column_letter
import os
import dill as pkl

from load_data import Context

def make_group_report(context, date, directory):
    wb_name = os.path.join(directory, f"PIList-{date}.xlsx")

    font = openpyxl.styles.Font(name="Times New Roman")
    fontb = openpyxl.styles.Font(name="Times New Roman", bold=True)
    def add_cell(sheet, row, col, val, font=font):
        cell = sheet.cell(row=row, column=col)
        cell.value = val
        cell.font = font

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "PI info"

    add_cell(sheet, 1, 1, "gid", font=fontb)
    add_cell(sheet, 1, 2, "projects", font=fontb)
    add_cell(sheet, 1, 3, "ngid", font=fontb)
    add_cell(sheet, 1, 4, "First name", font=fontb)
    add_cell(sheet, 1, 5, "Last name", font=fontb)
    add_cell(sheet, 1, 6, "Department", font=fontb)
    add_cell(sheet, 1, 7, "Email", font=fontb)
    for i in range(1, 7):
        sheet.column_dimensions[get_column_letter(i)].width = 20

    for n, group in context.groups.iterrows():
        n += 2
        gid = group[0]
        projects = group[1]
        last_name = group[2]
        first_name = group[3]
        email = group[4]
        department = group[5]
        ngid = group[6]

        add_cell(sheet, n, 1, gid)
        add_cell(sheet, n, 2, projects)
        add_cell(sheet, n, 3, ngid)
        add_cell(sheet, n, 4, first_name)
        add_cell(sheet, n, 5, last_name)
        add_cell(sheet, n, 6, department)
        add_cell(sheet, n, 7, email)

    wb.save(filename=wb_name)

if __name__ == "__main__":
    from datetime import date
    now = date.today()

    context = Context(verbosity = 0, insert_data = 0, path_to_quota = "/m31/reps/wekafs.qta", path_to_pkl = os.getenv("REPORT_DATA_PATH", os.getcwd()))
    make_group_report(context, now.strftime('%b%Y'), directory=os.getcwd())
