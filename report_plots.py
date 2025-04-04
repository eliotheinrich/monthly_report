import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter
import matplotlib as mpl

import numpy as np

MONTHS = {1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun", 7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"}

def date_label(date):
    d = str(date)
    return MONTHS[int(d[5:7])] + "-" + d[2:4]

plt.rcParams["font.size"] = 25

def plot_usage_by_group(context, group_usage, start_date, end_date, title=None, xlabel=None, threshold=-1, directory=None, output_extension="png", f=None):
    if directory is None:
        directory = ""
    else:
        if directory[-1] == "/":
            directory = directory[:-1]
    fig, ax = plt.subplots(figsize=(20,12))

    users, time = zip(*[(k, v) for k,v in group_usage.items() if v >= threshold])
    sorted_inds = np.flip(np.argsort(time))
    users, time = np.array(users)[sorted_inds], np.array(time)[sorted_inds]
    users = users[:threshold]
    time = time[:threshold]

    ax.barh([context.get_label(gid) for gid in users], time, color="orange", label="Requested")
        
    plt.yticks(rotation=0)
    ax.set_xlim(0, max(time)*1.2)
    if f is None:
        ax.xaxis.set_major_formatter(FuncFormatter(lambda x, p: str(int(x/1000)) + "k"))
    plt.title(f"{date_label(start_date)} - {date_label(end_date)}", y=1.0, pad=30)
    if title is not None:
        plt.suptitle(title)
    if xlabel is not None:
        plt.xlabel(xlabel)

    # Hide axis spines
    ax.spines["top"].set_visible(False)
    ax.spines["bottom"].set_visible(False)
    ax.spines["left"].set_visible(False)
    ax.spines["right"].set_visible(False)


    # Add gridlines
    for tick in ax.get_xticks():
        plt.axvline(tick, linewidth=0.25, color="k", zorder=0)

    # Hide ticks and remove bottom ytick corresponding to 0 hours
    ax.tick_params(axis="both", length=0)
    ax.set_xticks(ax.get_xticks()[1:])

    plt.savefig(f"{directory}/{title}.{output_extension}", bbox_inches="tight")
    plt.show()

def plot_yearly_usage(
        total_usage, 
        months,
        title: str = None,
        ylabel: str = None,
        max_usage: int = None,
        directory: str = None,
        output_extension="png",
        f=None
    ):

    if directory is None:
        directory = ""
    else:
        if directory[-1] == "/":
            directory = directory[:-1]

    num_months = len(months)
    start_date = months[0]
    end_date = months[-1]
    
    _, ax = plt.subplots(figsize=(12, 8))
    xticks = list(range(0, num_months))

    legend_items = []
    
    plt.plot(xticks, total_usage, color="orange")
    legend_items.append((mpl.patches.Rectangle([0,0],0,0, color="orange"), "Requested"))
    plt.fill_between(xticks, total_usage, color="orange")
 
    plt.suptitle(title)
    plt.title(f"{date_label(start_date)} - {date_label(end_date)}")


    # Format x-axis
    #ax.set_xlim(start_date, end_date)
    ax.set_xticks(xticks)
    ax.set_xticklabels([date_label(month) for month in months], rotation=30)


    # Show max usage in red
    if max_usage is None:
        max_usage = max(total_usage)*1.1
    #ax.axhline(max_usage, color="r", linestyle="--", alpha=0.5)

    # Format y-axis
    if ylabel is not None:
        plt.ylabel(ylabel)
    #ax.set_ylim(0, 1.2*max_usage)
    if f is None:
        ax.yaxis.set_major_formatter(FuncFormatter(lambda x, p: str(int(x/1000)) + "k"))

    # Hide axis spines
    ax.spines["top"].set_visible(False)
    ax.spines["bottom"].set_visible(False)
    ax.spines["left"].set_visible(False)
    ax.spines["right"].set_visible(False)


    # Add gridlines
    for tick in ax.get_yticks():
        plt.axhline(tick, linewidth=0.25, color="k", zorder=0)

    # Hide ticks and remove bottom ytick corresponding to 0 hours
    ax.tick_params(axis="both", length=0)
    ax.set_yticks(ax.get_yticks()[1:])

    plt.savefig(f"{directory}/{title}.{output_extension}", bbox_inches="tight")
    plt.show()

ABBREVIATIONS = {
    "Carroll School of Management": "CSOM", 
    "Lynch School of Education and Human Development": "LSOE",
    "Earth and Environmental Sciences": "EES",
}

def plot_usage_by_department(context, total_usage_by_group, start_date, end_date, title, directory=None, output_extension="png", f=None):
    if directory is None:
        directory = ""
    else:
        if directory[-1] == "/":
            directory = directory[:-1]
    usage_by_department = {}


    for gid, time in total_usage_by_group.items():
        department = context.get_department(gid)
        if department in ABBREVIATIONS:
            department = ABBREVIATIONS[department]
        if department not in usage_by_department:
            usage_by_department[department] = 0.
        usage_by_department[department] += time
    fig, ax = plt.subplots(figsize=(16,12))

    departments, time = zip(*[(k, v) for k,v in usage_by_department.items()])
    sorted_inds = np.flip(np.argsort(time))
    departments, time = np.array(departments)[sorted_inds], np.array(time)[sorted_inds]

    ax.barh(departments, time, color="orange")
    plt.yticks(rotation=0)
    plt.xlabel("CPU hours used")
    if f is None:
        ax.xaxis.set_major_formatter(FuncFormatter(lambda x, p: str(int(x/1000)) + "k"))

    plt.title(f"{date_label(start_date)} - {date_label(end_date)}", y=1.0, pad=30)
    if title is not None:
        plt.suptitle(title)

    # Hide axis spines
    ax.spines["top"].set_visible(False)
    ax.spines["bottom"].set_visible(False)
    ax.spines["left"].set_visible(False)
    ax.spines["right"].set_visible(False)


    # Add gridlines
    for tick in ax.get_xticks():
        plt.axvline(tick, linewidth=0.25, color="k", zorder=0)

    # Hide ticks and remove bottom ytick corresponding to 0 hours
    ax.tick_params(axis="both", length=0)

    plt.savefig(f"{directory}/{title}.{output_extension}", bbox_inches="tight")
    plt.show()

def plot_storage_by_department(context, storage_by_group, title=None, directory=None, output_extension="png", f=None):
    if directory is None:
        directory = ""
    else:
        if directory[-1] == "/":
            directory = directory[:-1]
    storage_by_department = {}

    for gid, storage_usage in storage_by_group.items():
        try:
            department = context.get_department(gid)
        except ValueError:
            continue

        if department in ABBREVIATIONS:
            department = ABBREVIATIONS[department]
        
        if department not in storage_by_department:
            storage_by_department[department] = 0.

        storage_by_department[department] += storage_usage
    fig, ax = plt.subplots(figsize=(16,12))

    departments, storage_usage = zip(*[(k, v) for k,v in storage_by_department.items()])
    sorted_inds = np.flip(np.argsort(storage_usage))
    departments, storage_usage = np.array(departments)[sorted_inds], np.array(storage_usage)[sorted_inds]

    ax.barh(departments, storage_usage, color="orange")
    plt.yticks(rotation=0)
    plt.xlabel("Storage space used (GB)")
    if f is None:
        ax.xaxis.set_major_formatter(FuncFormatter(lambda x, p: str(int(x/1000)) + "k"))

    if title is not None:
        plt.suptitle(title)

    # Hide axis spines
    ax.spines["top"].set_visible(False)
    ax.spines["bottom"].set_visible(False)
    ax.spines["left"].set_visible(False)
    ax.spines["right"].set_visible(False)


    # Add gridlines
    for tick in ax.get_xticks():
        plt.axvline(tick, linewidth=0.25, color="k", zorder=0)

    # Hide ticks and remove bottom ytick corresponding to 0 hours
    ax.tick_params(axis="both", length=0)

    plt.savefig(f"{directory}/{title.replace('/', '')}.{output_extension}", bbox_inches="tight")
    plt.show()


def plot_storage_by_group(context, storage_by_group, cutoff=None, title=None, directory=None, output_extension="png", f=None):
    if directory is None:
        directory = ""
    else:
        if directory[-1] == "/":
            directory = directory[:-1]
    fig, ax = plt.subplots(figsize=(16,12))

    groups, storage_usage = zip(*[(k, v) for k,v in storage_by_group.items()])
    sorted_inds = np.flip(np.argsort(storage_usage))
    groups, storage_usage = np.array(groups)[sorted_inds], np.array(storage_usage)[sorted_inds]
    groups = np.array([context.get_label(id) for id in groups])
    storage_used = sum(storage_usage)
    if cutoff is not None:
        groups, storage_usage = groups[:cutoff], storage_usage[:cutoff]
    
    displayed_storage = sum(storage_usage)
    
    if storage_used != displayed_storage:
        groups = np.append(groups, "Other")
        storage_usage = np.append(storage_usage, storage_used - displayed_storage)

    ax.barh(groups, storage_usage, color="orange")
    plt.yticks(rotation=0)
    plt.xlabel("Storage space used (GB)")
    if f is None:
        ax.xaxis.set_major_formatter(FuncFormatter(lambda x, p: str(int(x/1000)) + "k"))
    if title is not None:
        plt.title(title)

    # Hide axis spines
    ax.spines["top"].set_visible(False)
    ax.spines["bottom"].set_visible(False)
    ax.spines["left"].set_visible(False)
    ax.spines["right"].set_visible(False)


    # Add gridlines
    for tick in ax.get_xticks():
        plt.axvline(tick, linewidth=0.25, color="k", zorder=0)

    # Hide ticks and remove bottom ytick corresponding to 0 hours
    ax.tick_params(axis="both", length=0)

    plt.savefig(f"{directory}/{title.replace('/', '')}.{output_extension}", bbox_inches="tight")
    plt.show()

def plot_storage_by_group_piechart(context, storage_by_group, cutoff=None, total_storage=None, title=None, legend=True, directory=None, output_extension="png"):
    if directory is None:
        directory = ""
    else:
        if directory[-1] == "/":
            directory = directory[:-1]

    if legend:
        import matplotlib.gridspec as gridspec
        
        plt.figure(figsize=(18,10))
        gs = gridspec.GridSpec(1, 2, width_ratios=[2.5,1])
        ax = [plt.subplot(gs[0]), plt.subplot(gs[1])]
    else:
        _, ax = plt.subplots(figsize=(18,10))

    groups, storage_usage = zip(*[(k, v) for k,v in storage_by_group.items()])
    sorted_inds = np.flip(np.argsort(storage_usage))
    groups, storage_usage = np.array(groups)[sorted_inds], np.array(storage_usage)[sorted_inds]
    groups = np.array([context.get_label(id) for id in groups])
    storage_used = sum(storage_usage)
    if cutoff is not None:
        groups, storage_usage = groups[:cutoff], storage_usage[:cutoff]
    
    color_cycle = plt.rcParams["axes.prop_cycle"].by_key()["color"]
    colors = [color_cycle[i % len(color_cycle)] for i in range(len(groups))]
    
    displayed_storage = sum(storage_usage)
    
    if storage_used != displayed_storage:
        groups = np.append(groups, "Other")
        storage_usage = np.append(storage_usage, storage_used - displayed_storage)
        colors.append("0.6")

    if total_storage is None:
        total_storage = storage_used
    
    if total_storage - storage_used > 0:
        groups = np.append(groups, "Unused")
        storage_usage = np.append(storage_usage, total_storage - storage_used)
        colors.append("0.3")

    if legend:
        patches, _, _ = ax[0].pie(
            storage_usage, 
            autopct="%1.1f%%",
            colors=colors,
            startangle=90,
            pctdistance=1.2,
        )
        ax[1].legend(patches, groups, loc="center")
        ax[1].axis("off")
        ax = ax[0]
    else:
        ax.pie(
            storage_usage, 
            labels=groups, 
            autopct="%1.1f%%",
            colors=colors,
            startangle=90,
        )
    
    plt.yticks(rotation=0)
    ax.xaxis.set_major_formatter(FuncFormatter(lambda x, p: str(int(x/1000)) + "k"))
    if title is not None:
        plt.suptitle(title)

    # Hide axis spines
    ax.spines["top"].set_visible(False)
    ax.spines["bottom"].set_visible(False)
    ax.spines["left"].set_visible(False)
    ax.spines["right"].set_visible(False)


    # Add gridlines
    for tick in ax.get_xticks():
        plt.axvline(tick, linewidth=0.25, color="k", zorder=0)

    # Hide ticks and remove bottom ytick corresponding to 0 hours
    ax.tick_params(axis="both", length=0)

    plt.subplots_adjust(bottom=0., top=1.0)
    plt.savefig(f"{directory}/{title.replace('/', '')}.{output_extension}", bbox_inches="tight")
    plt.show()
    
# -- Generates excel spreadsheet with all usage information -- #

import openpyxl
from openpyxl.utils import get_column_letter

def monthyear(date):
    month = MONTHS[int(date[5:7])]
    year = date[:4]

    return month, year

def make_report_sheet(context, report, directory=None):
    if directory is None:
        directory = ""
    else:
        if directory[-1] == "/":
            directory = directory[:-1]
    months = report.months
    num_months = len(months)

    group_usage = report.get_group_usage()

    data_storage = report.query("homeStorage", idx=-1)
    scratch_storage = report.query("scratchStorage", idx=-1)
    project_storage = report.query("projectStorage", idx=-1)

    report_month, report_year = monthyear(months[-1])

    wb_name = f"{directory}/ClusterUsage{report_month}{report_year}.xlsx"

    wb = openpyxl.Workbook()

    font = openpyxl.styles.Font(name="Times New Roman")
    fontb = openpyxl.styles.Font(name="Times New Roman", bold=True)


    def make_sheet(sheet, usage_by_group):
        def add_cell(row, col, val, font=font):
            cell = sheet.cell(row=row, column=col)
            cell.value = val
            cell.font = font

        add_cell(4, 1, "Group", fontb)
        sheet.column_dimensions[get_column_letter(1)].width = 20
        add_cell(4, 2, "Department", fontb)
        sheet.column_dimensions[get_column_letter(2)].width = 20

        visited_years = set()

        for n,month in enumerate(months):
            sheet.column_dimensions[get_column_letter(3 + n)].width = 15
            month, year = monthyear(month)
            
            add_cell(5, 3 + n, month, fontb)
            if year not in visited_years:
                visited_years.add(year)
                add_cell(4, 3 + n, year, fontb)

        last_col = 4 + n
        add_cell(5, last_col, "1 Yr Total", fontb)
        sheet.column_dimensions[get_column_letter(last_col)].width = 20
        
        h = 5
        for n,gid in enumerate(context.gids):
            name = context.get_group_name(gid)
            department = context.get_department(gid)
            total_group_usage = sum([usage_by_group[gid][m] for m in range(num_months)])
            if total_group_usage < 1:
                continue

            h += 1
            add_cell(h, 1, name)
            add_cell(h, 2, department)
            for m,_ in enumerate(months):
                if usage_by_group[gid][m] > 1:
                    add_cell(h, 3 + m, round(usage_by_group[gid][m], 1))
            add_cell(h, last_col, round(total_group_usage, 1), fontb)

        last_row = h + 1
        add_cell(last_row, 1, "Total", fontb)

        total_usage = 0
        for i in range(num_months):
            monthly_usage = sum([usage_by_group[gid][i] for gid in context.gids])
            total_usage += monthly_usage
            add_cell(last_row, 3 + i, round(monthly_usage, 1), fontb)
        add_cell(last_row, last_col, round(total_usage, 1), fontb)

    def make_storage_sheet(sheet, storage):
        def add_cell(row, col, val, font=font):
            cell = sheet.cell(row=row, column=col)
            cell.value = val
            cell.font = font

        add_cell(4, 1, "Group", fontb)
        add_cell(4, 2, "Department", fontb)
        for n,label in enumerate(storage, 3):
            add_cell(4, n, label, fontb)
        
        last_col = 3 + len(storage)
        add_cell(4, last_col, "Total", fontb)

        for i in range(1, 3 + len(storage)):
            sheet.column_dimensions[get_column_letter(i)].width = 30

        h = 5
        for gid in context.gids:
            if not any([gid in storage[label] and storage[label][gid] > 1 for label in storage]):
                continue

            name = context.get_group_name(gid)
            department = context.get_department(gid)
            
            add_cell(h, 1, name)
            add_cell(h, 2, department)
            
            for n,label in enumerate(storage, 3):
                if gid in storage[label] and storage[label][gid] > 1:
                    add_cell(h, n, round(storage[label][gid]), 1)
                
            total_storage = sum([0 if gid not in storage[label] else storage[label][gid] for label in storage])
            add_cell(h, last_col, round(total_storage, 1), fontb)

            h += 1

    sheet = wb.active
    sheet.title = "CPU Usage"
    make_sheet(sheet, group_usage["cpuUsage"])

    sheet = wb.create_sheet("GPU Usage")
    make_sheet(sheet, group_usage["gpuUsage"])

    sheet = wb.create_sheet("Requested MEM")
    make_sheet(sheet, group_usage["reqMem"])

    sheet = wb.create_sheet("Storage")
    total_group_storage = {
        "/data/ storage (GB)": data_storage, 
        "/scratch/ storage (GB)": scratch_storage,
        "/project/ storage (GB)": project_storage,
    }

    make_storage_sheet(sheet, total_group_storage)

    if context.verbose:
        print(f"Saving {wb_name}")
    wb.save(filename=wb_name)
