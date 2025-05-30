import openpyxl
import dill as pkl
import os
from openpyxl.utils import get_column_letter

from load_data import Context 

def make_user_report(context, date, directory):
    wb_name = os.path.join(directory, f"UserList-{date}.xlsx")

    font = openpyxl.styles.Font(name="Times New Roman")
    fontb = openpyxl.styles.Font(name="Times New Roman", bold=True)

    def add_cell(sheet, row, col, val, font=font):
        cell = sheet.cell(row=row, column=col)
        cell.value = val
        cell.font = font

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "User info"
    add_cell(sheet, 1, 1, "uid", font=fontb)
    add_cell(sheet, 1, 2, "projects", font=fontb)
    add_cell(sheet, 1, 3, "gid", font=fontb)
    add_cell(sheet, 1, 4, "nuid", font=fontb)
    add_cell(sheet, 1, 5, "First name", font=fontb)
    add_cell(sheet, 1, 6, "Last name", font=fontb)
    add_cell(sheet, 1, 7, "Email", font=fontb)
    for i in range(1, 7):
        sheet.column_dimensions[get_column_letter(i)].width = 20

    for n, user in context.users.iterrows():
        n += 2
        uid = user[0]
        nuid = user[1]
        projects = user[2]
        gid = user[3]
        first_name = user[4]
        last_name = user[5]
        email = user[6]

        add_cell(sheet, n, 1, uid)
        add_cell(sheet, n, 2, projects)
        add_cell(sheet, n, 3, gid)
        add_cell(sheet, n, 4, nuid)
        add_cell(sheet, n, 5, first_name)
        add_cell(sheet, n, 6, last_name)
        add_cell(sheet, n, 7, email)

    wb.save(filename=wb_name)


if __name__ == "__main__":
    from datetime import date
    now = date.today()

    context = Context(verbosity = 0, insert_data = 0, path_to_quota = "/m31/reps/wekafs.qta", path_to_pkl = os.getenv("REPORT_DATA_PATH", os.getcwd()))
    make_user_report(context, now.strftime("%b%Y"), directory=os.getcwd())
